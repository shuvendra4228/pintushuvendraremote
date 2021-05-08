import openpyxl

book = openpyxl.load_workbook(path)
sheet = book.active
dicty = dict()
# cell = sheet.cell(row=1,column=1),row and both starts with 1 not with 0
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(1, sheet.max_column + 1):
            # print(sheet.cell(row=i, column=j).value)
            dicty[sheet.cell(row=i, column=1).value] = sheet.cell(row=i, column=j).value
print(dicty)
