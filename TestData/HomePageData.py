import openpyxl


# per each page we derive testdata for that specific page
class HomePageData:
    #test_HomePageData = [{"firstname": "rahul", "lastname": "shetty", "gender": "male"},{"firstname": "mona", "lastname": "lisa", "gender": "female"},{"firstname": "asu", "lastname": "tosh", "gender": "male"}]

    @staticmethod # to avoid object creation
    def getTestData(testcase_Name):
        book = openpyxl.load_workbook(path)
        sheet = book.active
        dicty = dict()
        # cell = sheet.cell(row=1,column=1),row and both starts with 1 not with 0
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_Name:
                for j in range(1, sheet.max_column + 1):
                    # print(sheet.cell(row=i, column=j).value)
                    dicty[sheet.cell(row=i, column=1).value] = sheet.cell(row=i, column=j).value
# if some fields are blank or some extra field are there for a particular testcase
# then in test_HomePage we use the key as per requirement and get the value
        return [dicty]